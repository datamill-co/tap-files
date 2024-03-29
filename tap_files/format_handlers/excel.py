import re

from tap_files.format_handlers.base import BaseFormatHandler
from tap_files.discover_utils import SDC_SOURCE_LINENO_COLUMN

def get_gen(workbook, sheet_name):
    def get_stream():
        return workbook.get_sheet_by_name(sheet_name).values
    return get_stream

def normalize_name(name):
    return re.sub(r'[^a-z0-9_]', '_', name.lower())

class ExcelFormatHandler(BaseFormatHandler):
    format_name = 'excel'
    extensions = ['xlsx', 'xls']
    default_extension = 'xlsx'
    file_mode = 'rb'

    def _get_rows_reader(self, stream_name, stream_config, ext, file, reader_gen=None):
        format_options = stream_config.get('format_options', {})
        fieldnames = format_options.get('fieldnames')
        tiered_headers = format_options.get('tiered_headers')
        preheader_skip_lines = format_options.get('preheader_skip_lines')
        findheader = format_options.get('findheader')
        skip_lines = format_options.get('skip_lines')

        if reader_gen:
            reader = reader_gen()
        else:
            reader = self._get_excel_reader(ext, format_options, file)

        if findheader and not preheader_skip_lines:
            cur_cols = 0
            max_cols = 0
            last_cols = 0
            header_row_num = 0
            i = 0
            stable = 0
            for row in reader:
                for k in range(len(row)):
                    if row[-(k+1)] is not None:
                        cur_cols = len(row) - k
                        break

                if last_cols == cur_cols:
                    stable += 1
                else:
                    stable = 0
                    max_cols = 0

                if cur_cols > max_cols:
                    max_cols = cur_cols
                    header_row_num = i

                last_cols = cur_cols
                i += 1

                if stable > 10:
                    preheader_skip_lines = header_row_num # header number is zero indexed
                    if preheader_skip_lines <= 0:
                        preheader_skip_lines = None
                    break

            if stable < 10:
                raise Exception('Header finder could not find a stable header row for: {}'.format(stream_name))

            # reset reader
            if reader_gen:
                reader = reader_gen()
            else:
                reader = self._get_excel_reader(ext, format_options, file)

        line_num = 0

        if preheader_skip_lines:
            for i in range(preheader_skip_lines):
                next(reader)
                line_num += 1

        if fieldnames:
            headers = fieldnames
        else:
            raw_headers = next(reader)
            if tiered_headers:
                secondary_headers = next(reader)
            headers = []
            for i in range(len(raw_headers)):
                header = raw_headers[i]
                if not tiered_headers and (header is None or header.strip() == ''):
                    headers.append(None) ## kept so we know to skip below
                else:
                    header_value = str(header or '').strip().replace('\n', ' ').replace("'", '')
                    if tiered_headers:
                        header_value += ' ' + str(secondary_headers[i] or '').strip().replace('\n', ' ')
                    
                    headers.append(header_value)
            line_num = 1 # header is 1

        if skip_lines:
            for i in range(skip_lines):
                next(reader)
                line_num += 1

        for row in reader:
            record = {}
            for i in range(len(headers)):
                header = headers[i]
                if header is not None:
                    record[header] = row[i]
            line_num += 1
            record[SDC_SOURCE_LINENO_COLUMN] = line_num

            yield record

    def _get_excel_reader(self, ext, format_options, file):
        if ext == 'xlsx':
            return self._xlsx(format_options, file)
        elif ext == 'xls':
            return self._xls(format_options, file)
        else:
            raise Exception('Excel extension "{}" not supported'.format(ext))

    def _get_streams(self, stream_config, file):
        from openpyxl import load_workbook

        format_options = stream_config.get('format_options', {})
        skip_by_index = format_options.get('skip_worksheet_by_index', [])
        skip_by_name = format_options.get('skip_worksheet_by_name', [])

        workbook = load_workbook(file)

        streams = []
        i = 0
        for sheet_name in workbook.get_sheet_names():
            k = i
            i += 1
            if k in skip_by_index or sheet_name in skip_by_name:
                continue

            streams.append([
                stream_config['stream_name'] + '_' + normalize_name(sheet_name),
                get_gen(workbook, sheet_name)
            ])

        return streams

    def _xlsx(self, format_options, file):
        from openpyxl import load_workbook

        workbook = load_workbook(file)

        worksheet_name = format_options.get('worksheet_name')
        worksheet_index = format_options.get('worksheet_index')
        if worksheet_name:
            worksheet = workbook.get_sheet_by_name(format_options['worksheet_name'])
        elif worksheet_index:
            worksheet = workbook.worksheets[worksheet_index]
        else:
            worksheet = workbook.worksheets[0]

        return worksheet.values

    def _xls(self, format_options, file):
        import xlrd

        workbook = xlrd.open_workbook(on_demand=True, file_contents=file.read())

        worksheet_name = format_options.get('worksheet_name')
        if worksheet_name:
            worksheet = workbook.sheet_by_name(format_options['worksheet_name'])
        else:
            worksheet = workbook.sheet_by_index(0)

        for row in worksheet.get_rows():
            out_row = []
            for cell in row:
                out_row.append(cell.value)
            yield out_row
